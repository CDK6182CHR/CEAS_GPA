"""
成绩有效性处理器。
聚合于Calculator中
2020.08.24添加说明：
grade.note 用于处理本系统的映射。但体育、英语的正常映射不添加。
grade.note2  教务系统原始备注。
grade.flag  教务系统标记。
以上三项都加到最后备注里。
"""
from .gradelib import GradeLib
from .studentgrade import StudentGrade,Grade
from .courselib import CourseLib,Course
from .studentlist import StudentList,Student
from .coursemaps import multiMapTable,singleMapTable,PENumbers,PERegex,substituteMapTable
from datetime import datetime
import openpyxl

import re


class GradeValidator:
    LogDir = '../log/'
    def __init__(self,gradeLib,courseLib:CourseLib,mode,start,end,useSubstitute,useFirst,firstYear:str,
                 zeroForAbsent,studentList:StudentList):
        self.gradeLib = gradeLib  # type:GradeLib
        self.courseLib = courseLib  # type:CourseLib
        self.mode = mode  # type:int
        self.startSemester = start  # type:int
        self.endSemester = end  # type:int
        self.useSubstitute = useSubstitute  # type:bool
        self.useFirst = useFirst  # type:bool
        self.firstYear = firstYear  # type:str  # 入学年份字符串。用于做学期映射。
        self.zeroForAbsent = zeroForAbsent  # type:bool
        self.studentList = studentList  # type:StudentList
        self.transStudentList = StudentList()  # type:StudentList
        self.readTransList()
        self.log = open(
            self.LogDir+datetime.now().strftime('log %Y-%m-%d_%H-%M-%S.txt'),
            'w',encoding='utf-8',errors='ignore'
        )
        self.log.write(f"mode {self.mode}\n")

    def __del__(self):
        pass
        # self.log.close()

    def validate(self):
        """
        接口方法
        """
        self.removeDuplicate()
        self.preMapProcess()
        self.multiMap()
        self.singleMap()
        self.choose()
        self.PEMap()
        if self.useSubstitute:
            self.substituteMap()
        self.postMapProcess()
        self.filtByList()
        self.postFiltProcess()

    def readTransList(self,filename='../data/跨大类.xlsx'):
        """
        读取转专业名单。名单中的，缺课不按0分处理。
        """
        wb = openpyxl.load_workbook(filename)
        ws = wb.get_sheet_by_name(wb.sheetnames[0])
        for row in range(1,ws.max_row+1):
            try:
                num = int(ws.cell(row,1).value)
                name = ws.cell(row,2).value
            except:
                print(f"Validator::ReadTransList  数据错误！行号{row}")
            else:
                s = self.studentList.studentByNumber(num)
                if s is not None:
                    self.transStudentList[num] = s
                else:
                    print(f"Validator::ReadTransList  Invalid number: {num}")

    def multiMap(self):
        """
        对体验英语的特殊处理。一门体验英语映射到同一学期的两门英语课中。
        仍在list中完成。
        【注意学分是按照映射后的课程！】
        """
        self.log.write("**********multiMap**********\n")
        for stu_number,stu_grade in self.gradeLib.items():
            stu_grade:StudentGrade
            for courseId,grade in stu_grade._source.copy().items():
                if multiMapTable.get(courseId,None) is None:
                    continue
                # 在multiMap映射表中
                if len(grade)!=1:
                    print(f"multiMap: more than one courses to be multi map, student {stu_number} course {courseId} length {len(grade)}")

                txt = str(grade[0])
                oldName = grade[0].course_name

                for newId in multiMapTable[courseId]:
                    mappedCourse = self.courseLib.courseById(newId)
                    if mappedCourse is None:
                        print(f"Unexpceted None course {newId},"
                              f"mapped from {courseId}")
                        continue
                    grade1 = stu_grade.addGrade(
                        Grade(
                            stu_number,
                            newId,
                            mappedCourse.name,
                            mappedCourse.credits if newId != '00010011C' else 0,  # 特别处理二层次数学问题
                            grade[0].semester,
                            grade[0].grade_type,
                            grade[0].total,
                            grade[0].note2,
                            grade[0].flag
                        )
                    )
                    grade1.note += f"mapped: {oldName} {grade1.credits}学分"
                    grade1.oldId = courseId
                self.log.write(f"Multimap: {txt}->{grade1}\n")
                del stu_grade._source[courseId]

    def singleMap(self):
        """
        处理英语课名称不同需要映射的情况。需考虑多条数据。
        """
        self.log.write("**********singleMap*********\n")
        for stu_number,stu_grade in self.gradeLib.items():
            stu_grade:StudentGrade
            for courseId, grades in stu_grade._source.copy().items():
                newId = singleMapTable.get(courseId,None)
                if newId is None:
                    continue
                mappedCourse = self.courseLib.courseById(newId)
                if mappedCourse is None:
                    self.log.write(f"目标课程不在列表中{newId} mapped from: {courseId}")
                    continue
                for grade in grades:
                    old_descrip = str(grade)
                    grade.course_name = mappedCourse.name
                    grade.course_number = mappedCourse.id
                    # grade.note+=" mapped"
                    self.log.write(f"{old_descrip}->{mappedCourse}\n")
                stu_grade._source[newId]=stu_grade._source[courseId]
                del stu_grade._source[courseId]

    def removeDuplicate(self):
        """
        删除层次上冲突的课程。
        """
        self.log.write("*******removeDuplicate******\n")
        for stu_num, stu_grade in self.gradeLib.items():
            stu_grade:StudentGrade
            if stu_grade._source.get('00010014B') is not None:  # 有体验数学下就推定按照体验数学
                for course_id in ('00010011A','00010011B','00010011C'):
                    try:
                        del stu_grade._source[course_id]
                    except KeyError:
                        pass
                    else:
                        self.log.write(f"删除一层次数学，按体验数学处理：{stu_num} delete {course_id}\n")

    def preMapProcess(self):
        """
        放在所有map前。
        特殊处理所有的需要调整分数的情况。
        """
        self.log.write("*******preMapProcess******\n")
        for num,stu_grade in self.gradeLib.items():
            stu_grade:StudentGrade
            for courseId, grades in stu_grade._source.items():
                for grade in grades:
                    if grade.course_number.startswith('00010014'):
                        # 体验数学
                        tot = grade.total
                        grade.total*=0.6
                        grade.note2 += f' 体验数学0.6倍率。原：{tot}'
                        self.log.write(f'体验数学按0.6倍率计算：{grade}\n')
                    elif grade.course_number.startswith('41000110'):
                        tot = grade.total
                        grade.total*=0.8
                        grade.note2 += f' 体验英语0.8倍率。原：{tot}'
                        self.log.write(f'体验英语按0.8倍率计算：{grade}\n')

    def postMapProcess(self):
        """
        所有映射结束后特殊处理，用于处理体验数学下的问题。
        """
        self.log.write("*******postMapProcess******\n")
        for num,stu_grade in self.gradeLib.items():
            stu_grade:StudentGrade
            for courseId, grades in stu_grade._source.items():
                for grade in grades:
                    if grade.course_number == '00010011B' and grade.oldId == '00010014B':
                        # 体验数学下，6个学分
                        grade.credits = 6
                        self.log.write(f'体验数学下按照6学分计算：{grade}\n')

    def postFiltProcess(self):
        """
        挑选完必修课后，再来特殊处理
        用于处理：不按*层次计算，但多了高等代数1的问题。
        """
        self.log.write("*******postFiltProcess******\n")
        psudoIds = ['11000020A','11000030']  # 赝课程的ID
        levelIds = ['30000010A','11000010A']
        for stu_num,stu_grade in self.gradeLib.copy().items():
            stu_grade:StudentGrade
            for id in psudoIds:
                g = stu_grade.getCourseGradeObject(id)
                g0 = stu_grade.getCourseGradeObject('00010011A')
                if g is not None and (g0.mappedFlag != Grade.SubstituteMapped or g0.oldId not in levelIds):
                    del stu_grade._table[id]
                    print(f'非*层次数学，删除赝课程：{g}')
                    self.log.write(f'非*层次数学，删除赝课程：{g}\n')

    def choose(self):
        """
        在有多条记录的选项中，选择一条记录来处理
        """
        self.log.write("*******choose******\n")
        for stu_number,stu_grade in self.gradeLib.items():
            stu_grade:StudentGrade
            for courseId, grades in stu_grade._source.copy().items():
                # 2020.08.25添加：无条件剔除“注销”成绩
                if True:
                    if len(grades) == 1:
                        continue
                    while len(grades) > 1:
                        for grade in grades.copy():
                            if '注销' in grade.grade_type:
                                print(f'By type 注销: delete {grade}')
                                self.log.write(f'By type 注销: delete {grade}\n')
                                grades.remove(grade)
                                break
                        else:
                            break
                if self.useFirst:
                    if len(grades)==1:
                        continue
                    # 先按学期排序
                    grades.sort(key=lambda grade:grade.semester)
                    for grade in grades.copy():
                        if grade.semester != grades[0].semester:
                            grades.remove(grade)
                            self.log.write(f"By semester: delete {grade}\n")
                    if len(grades) == 1:
                        continue
                    # 一个学期内应该最多两条记录，一个原始一个补考
                    # 现在只剩两个
                    for grade in grades.copy():
                        if '补考' in grade.flag or '重修' in grade.flag:
                            grades.remove(grade)
                            self.log.write(f"By flag {grade.flag}: delete {grade}\n")
                    if len(grades) >= 2:
                        print(f"choose: unknown multi courses length {len(grades)}",*grades)
                else:  # !useFirst
                    # 只选择type为“正式”的。
                    if len(grades) == 1:
                        continue
                    for grade in grades.copy():
                        if grade.grade_type == '无效':
                            grades.remove(grade)
                            self.log.write(f"By type {grade.grade_type}: delete {grade}\n")
                            if len(grades) == 1:
                                break
                    if len(grades) == 1:
                        continue
                    for grade in grades.copy():
                        if grade.grade_type and grade.grade_type != '正式':
                            grades.remove(grade)
                            self.log.write(f"By type {grade.grade_type}: delete {grade}\n")
                            if len(grades) == 1:
                                break
                    if len(grades) != 1:
                        print(f"Unexpected length of grades after choose: "
                              f"{len(grades)} {courseId} {stu_grade.student}")
            stu_grade._ok = True

    def filtByList(self):
        """
        按所选模式筛选课程，并放进table中
        """
        self.log.write("**********filtByList***********\n")
        for stu_number,stu_grade in self.gradeLib.items():
            stu_grade:StudentGrade
            student = stu_grade.student
            major = stu_grade.student.major
            for course_id, course in self.courseLib.items():
                course:Course
                if not self.startSemester <= course.semester <= self.endSemester:
                    continue
                if course.majorTypes[major] & self.mode:
                    # 添加这门课程！
                    grades = stu_grade._source.get(course_id,None)
                    if grades is None:
                        stu_grade.absentCount += 1
                        stu_grade.absentNames.append(course.name)
                        if self.zeroForAbsent and course.credits:  # for EEC: 非零学分才替代。0学分是菜单课。
                            # 2020.08.25：删除跨大类不用0分替代的逻辑。
                            # if self.transStudentList.get(stu_grade.student.number,None) is not None:
                            #     self.log.write(f"跨大类学生缺少课程：{course} {student} 按忽略处理")
                            # else:
                            stu_grade._table[course_id]=Grade(
                                student.number,course_id,course.name,
                                course.credits,'NA','NA',0.0,'缺数据',''
                            )
                            self.log.write(f"缺少必修课程：{course} {student} 自动添加0分数据\n")
                            # print(f"缺少必修课程：{course} {student} 自动添加0分数据")
                    else:
                        grade = grades[0]
                        stu_grade._table[course_id]=grade
                        if grade.course_name != course.name:
                            print(f"课程名不一致 {course} {grade}")
                            # self.log.write(
                            #     f"课程名不一致 {course} {grade}\n")
                        elif grade.credits != course.credits:
                            if course.majorTypes[student.major] == 3:
                                # 2020.08.25新增：只对3类课程打印学分数不一致警告。1类课程是菜单课，不管
                                print(f"学分数不一致 {course} {grade}")
                                self.log.write(
                                    f"学分数不一致 {course} {grade}\n")
                                grade.note2 += f" {grade.credits}学分"
                            elif course.majorTypes[student.major] == 1: # 对于菜单课的特殊处理，按照0学分！
                                grade.credits = course.credits

                        # 2019.10.26删除此逻辑
                        # if '补考' in grade.flag and grade.total >= 60:
                        #     self.log.write(
                        #         f"补考及格按60分处理：{grade}\n"
                        #     )
                        #     grade.total = 60
                        del stu_grade._source[course_id]
            for course_id,grades in stu_grade._source.items():
                self.log.write(f"非必修课 {grades[0]}\n")

    def PEMap(self):
        """
        体育课映射。按学期，优先映射到正常修读的学期对应的课程号。
        precondition: 各个课程号只有一条有效数据。
        """
        self.log.write("*************PEMap*************\n")
        for stu_number,stu_grade in self.gradeLib.items():
            stu_grade:StudentGrade
            highers = []  # 高年级修读的课程
            for course_id,grades in stu_grade._source.copy().items():
                grade = grades[0]
                if not re.match(PERegex,grade.course_number):
                    continue
                seme = grade.mappedSemester(self.firstYear)
                if seme>4:
                    highers.append(grade)
                    continue
                newId = PENumbers[seme-1]
                if self.courseLib.courseById(newId) is None:
                    self.log.write(f"目标课程不在列表中：{newId}\n")
                    continue
                self.log.write(f"{grade} -> {newId}\n")
                note = f' mapped from {grade}'
                grade.course_number = newId
                grade.course_name = self.courseLib.courseById(newId).name
                grade.note+=note
                del stu_grade._source[course_id]
                stu_grade.addGrade(grade)
            for grade in highers:
                for id in PENumbers:
                    if stu_grade._source.get(id,None) is not None:
                        continue
                    # 补充体育课数据
                    oldId = grade.course_number
                    grade.course_number = id
                    self.log.write(f"high grade PE map {grade}->{id}\n")
                    grade.course_name = self.courseLib.courseById(id).name
                    del stu_grade._source[oldId]
                    stu_grade.addGrade(grade)
                    break
                else:
                    print(f"Overmuch PE courses: {grade}")
                    self.log.write(f"Overmuch PE courses: {grade}\n")

    def substituteMap(self):
        """
        替代课程映射。
        preconditon: 每个课程号下有且仅有一个
        """
        self.log.write(f"*************substituteMap*************\n")
        for stu_number,stu_grade in self.gradeLib.items():
            stu_grade:StudentGrade
            for course_id,grades in stu_grade._source.copy().items():
                grade = grades[0]
                oldId = grade.course_number
                newId = substituteMapTable.get(grade.course_number,None)
                if newId is None:
                    continue
                # 找到替代课程，检查是否存在标准课程
                if stu_grade._source.get(newId,None) is not None:
                    self.log.write(f"Standard course existed, wont map: {newId},{grade}\n")
                    continue
                # 开始映射
                course = self.courseLib.courseById(newId)
                if course is None:
                    print(f"Mapped course not found: {newId}, from {grade}")
                    exit(343)
                self.log.write(f"substitute {grade} -> {course}\n")
                grade.note += f"mapped from {grade}"
                grade.course_number = course.id
                grade.course_name = course.name
                grade.oldId = oldId
                grade.mappedFlag = Grade.SubstituteMapped
                del stu_grade._source[course_id]
                stu_grade.addGrade(grade)



