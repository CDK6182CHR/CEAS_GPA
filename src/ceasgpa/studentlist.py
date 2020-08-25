"""
学生信息类，按学号->Student映射对象布置。
"""
import json
from .student import Student
class StudentList(dict):
    def __init__(self):
        super(StudentList, self).__init__()

    NumberCol=1
    NameCol=2
    MajorCol=3
    StartRow=2
    def readList(self,filename='../data/stu_list.xlsx'):
        self.clear()
        import openpyxl
        wb = openpyxl.load_workbook(filename,read_only=True)
        ws = wb.get_sheet_by_name(wb.sheetnames[0])
        for row in range(self.StartRow,ws.max_row+1):
            n_str = ws.cell(row,self.NumberCol).value
            try:
                n=int(n_str)
            except ValueError:
                print("Invalid student number at row",row,n_str)
                n=0
            name = ws.cell(row,self.NameCol).value
            major = ws.cell(row,self.MajorCol).value
            self[n]=Student(name,n,major)
        print(f"StudentList::Read len={len(self)}")
        wb.close()

    def saveJson(self,filename='../data/stu_list.json'):
        fp = open(filename,'w',encoding='utf-8',errors='ignore')
        data = {}
        for number,stu in self.items():
            data[number]=stu.__dict__
        json.dump(data,fp,ensure_ascii=False)
        fp.close()

    def readJson(self,filename='../data/stu_list.json'):
        fp = open(filename,encoding='utf-8',errors='ignore')
        dt = json.load(fp)
        for n,st in dt.items():
            self[int(n)]=Student(**st)

    def studentByNumber(self,number:int)->Student:
        return self[number]