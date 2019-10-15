import openpyxl,re

multiMapTable = {
    "41000110A":[
        "Read1",
        "Listen1"
    ],
    "41000110B":[
        "Read2",
        "Listen2"
    ]
}

singleMapTable = {
    # 第一学期，倒数第二位2->读写，3->听说，最后一位是层次
    "00020021A":"Read1",
    "00020022A":"Read1",
    "00020023A":"Read1",
    "00020031A":"Listen1",
    "00020032A":"Listen1",
    "00020033A":"Listen1",
    # 第二学期一、三层次常规课
    "00020021B":"Read2",
    "00020031B":"Listen2",
    "00020023B":"Read2",
    "00020033B":"Listen2",
    # 听说二层
    "00020032B1":"Listen2",  # 高级英语口语
    "00020032B2":"Listen2",  # 视听说与跨文化交际
    "00020032B3":"Listen2",  # 国际学术交流英语
    "00020032B4":"Listen2",  # 英文电影视听说与文化赏析
    # 读写二层
    "00020022B1":"Read2",  # 新闻英语读写
    "00020022B2":"Read2",  # 翻译基础与实践
    "00020022B3":"Read2",  # 批判性阅读：美国社会与文化
    "00020022B5":"Read2",  # 学术性阅读：语言与思维
    "00020022B6":"Read2",  # 学术英语写作
}

PERegex = re.compile(r'00040.*?')
PENumbers = (
    '00040010A',
    '00040010B',
    '00040010C',
    '00040010D',
)

wb = openpyxl.load_workbook('../data/map/substitute.xlsx')
ws = wb.get_sheet_by_name(wb.sheetnames[0])

substituteMapTable = {}

# openpyxl, 索引从1开始
StartRow = 4
OriginCol = 1
NewCol = 3

for row in range(StartRow,ws.max_row+1):
    originNumber = ws.cell(row,OriginCol).value
    newNumber = ws.cell(row,NewCol).value
    if originNumber:
        substituteMapTable[originNumber]=newNumber
