"""
课程库类。继承dict，数据结构为：课程名->Course对象。
在此处实现IO。
暂时用课程名映射。
"""
from typing import Union,Dict
from .course import Course
from datetime import datetime
import json

class CourseLib(dict):
    ValidMajors = ('材料物理','材料化学','光电信息','新能源','生物医学工程')
    def __init__(self):
        super(CourseLib, self).__init__()

    HeaderRow = 3
    CourseIdCol = 1
    CourseNameCol = 2
    CreditCol = 3
    SemesterCol = 4
    MajorStartCol = 6  # 个数为ValidMajors
    def readLibFile(self,filename='../data/课程性质数据维护.xlsx'):
        import openpyxl
        self.clear()
        wb = openpyxl.load_workbook(filename,read_only=True)
        ws = wb.get_sheet_by_name(wb.sheetnames[0])
        major_cols = {}  # type: Dict[int,str]
        # 先读取表格中的列号对应专业名称
        for col in range(CourseLib.MajorStartCol,
                         CourseLib.MajorStartCol+len(CourseLib.ValidMajors)):
            majorName = ws.cell(CourseLib.HeaderRow,col).value.strip()
            if majorName not in CourseLib.ValidMajors:
                print(f"Warning: Invalid major name {majorName} at column {col}")
            major_cols[col]=majorName
        # 逐行读取数据
        for row in range(CourseLib.HeaderRow+1,ws.max_row+1):
            courseName = ws.cell(row,CourseLib.CourseNameCol).value
            courseId = ws.cell(row,CourseLib.CourseIdCol).value  # 可能是None
            if courseId is None:
                courseId=''
            cre_str = ws.cell(row,CourseLib.CreditCol).value
            try:
                cre = int(cre_str)
            except ValueError:
                print("Invalid credit value at row",row,cre_str)
                cre=-1
            seme_str = ws.cell(row,CourseLib.SemesterCol).value
            try:
                seme = int(seme_str)
            except ValueError:
                print("Invalid semester value at row",row,seme_str)
                seme = 0
            course = Course(courseName,courseId,seme,cre)
            for col in range(CourseLib.MajorStartCol,
                             CourseLib.MajorStartCol+len(CourseLib.ValidMajors)):
                tp_raw = ws.cell(row,col).value
                try:
                    tp=int(tp_raw)
                except:
                    # 不填的默认为0
                    tp=0b00
                course.addMajor(major_cols[col],tp)
            # self[courseName]=course
            self[courseId]=course
        wb.close()

    def saveJson(self,filename='../data/courseLib.json'):
        data = [
            course.__dict__ for course in self.values()
        ]
        with open(filename,'w',encoding='utf-8',errors='ignore') as fp:
            json.dump(data,fp,ensure_ascii=False)
        with open(filename.rstrip('on'),'w',encoding='utf-8',errors='ignore') as fp:
            fp.write("var courseLib=")
            fp.write(json.dumps(data,ensure_ascii=False))
            fp.write(';\n')

    def saveMarkdown(self,filename='../data/分专业课程清单.md'):
        """
        按照专业->类别->学期的顺序计算。
        """
        fp = open(filename,'w',encoding='utf-8',errors='ignore')
        fp.write('# 现工院课程学分绩计算清单\n')
        for major in self.ValidMajors:
            fp.write(f'## {major}\n')
            for tpCode,tpName in Course.Types.items():
                fp.write(f'### {tpName}\n')
                for seme in range(1,9):
                    fp.write(f'#### 第{seme}学期\n')
                    fp.write("课程号 | 课程名 | 学分 \n")
                    fp.write("-|-|-\n")
                    for course in self.values():
                        if course.getMajorType(major) & tpCode and course.semester==seme:
                            fp.write(f'{course.id} | {course.name} | {course.credits} | \n')
        fp.close()

    def saveCourseMarkdown(self,filename='../data/课程清单.md'):
        """
        以课程为主要字段计算。按学期排序，同一学期内按课程号排序。
        """
        fp = open(filename, 'w', encoding='utf-8', errors='ignore')
        fp.write('# 现工院学分绩课程一览表\n')
        fp.write(f'更新时间：{datetime.now().strftime("%Y-%m-%d %H-%M-%S")}\n')
        fp.write('学期 | 课程号 | 课程名 | '+' | '.join(CourseLib.ValidMajors)+'\n')
        fp.write('-|-|-|'+'-|'*len(CourseLib.ValidMajors)+'\n')
        lst = list(self.values())
        lst.sort(key=lambda x:(x.semester,x.id))
        for course in lst:
            course:Course
            fp.write(f"{course.semester} | {course.id} | {course.name} |")
            for major in CourseLib.ValidMajors:
                fp.write(CourseLib.getMajorStr(course,major)+' | ')
            fp.write('\n')
        fp.close()


    @staticmethod
    def getMajorStr(course:Course,major:str):
        s=[]
        for tpCode,tpName in Course.BriefTypes.items():
            if course.majorTypes[major] & tpCode:
                s.append(tpName)
        if s:
            return ', '.join(s)
        return '—'

    def readJson(self,filename='../data/courseLib.json'):
        with open(filename,encoding='utf-8',errors='ignore') as fp:
            data = json.load(fp)
            for st in data:
                # self[st['name']]=Course(**st)
                self[st['id']]=Course(**st)

    def courseByName(self,name)->Course:
        return self[name]

    def courseById(self,id)->Course:
        for _,course in self.items():
            if course.id == id:
                return course
        return None

    def majorCourseList(self,major:str,mode:int,start,end)->list:
        lst = []
        for id,course in self.items():
            if course.getMajorType(major) & mode and start <= course.semester <= end:
                lst.append(course)
        return lst

