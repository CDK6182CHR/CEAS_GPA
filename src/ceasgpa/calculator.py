"""
计算器主类
"""
from .courselib import CourseLib, Course
from .gradelib import GradeLib,StudentGrade
from .studentlist import StudentList
from .validator import GradeValidator
import openpyxl
from openpyxl.comments import Comment
from datetime import datetime

class GpaCalculator:
    def __init__(self,mode,start,end,useSubstitute,useFirst,firstYear,zeroForAbsent,precision):
        self.courseLib = CourseLib()
        self.courseLib.readLibFile()
        self.studentList = StudentList()
        self.studentList.readList()
        self.gradeLib = GradeLib(self.studentList)
        self.gradeLib.readFileList()
        self.precision = precision
        self.mode = mode
        self.start = start
        self.end = end
        self.useSubstitute = useSubstitute
        self.useFirst = useFirst
        self.firstYear = firstYear
        self.zeroForAbsent = zeroForAbsent
        self.validator = GradeValidator(self.gradeLib,self.courseLib,mode,start,end,
                                        useSubstitute,useFirst,firstYear,zeroForAbsent,self.studentList)

    def validate(self):
        """
        筛选和整理数据
        """
        self.validator.validate()

    def calculate(self):
        for stu_number,stu_grade in self.gradeLib.items():
            print(stu_grade.student,stu_grade.calculate())

    def saveExcel(self,filename='../data/output.xlsx'):
        wb = openpyxl.Workbook()
        ws = wb.active
        if True:
            ws.title = '配置'
            """
            mode = 2
            startSemester = 1
            endSemester = 5
            useSubstitute = 1
            useFirst = 1
            firstYear = 2017
            zeroForAbsent = 1
            precision = 4
            """
            ws.append(('此文件由CEAS GPA Calculator自动生成，直接修改本文件可能会被后续运行覆盖。',))
            ws.append(('项目','配置'))
            ws.append(('课程集','1-学年' if self.mode==1 else '2-保研'))
            ws.append(('开始学期',self.start))
            ws.append(('结束学期',self.end))
            ws.append(('允许替代课程','是' if self.useSubstitute else '否'))
            ws.append(('仅使用初次成绩','是' if self.useFirst else '否'))
            ws.append(('当前计算年级',self.firstYear))
            ws.append(('缺课使用0分替代','是' if self.zeroForAbsent else '否'))
            ws.append(('输出小数位数',self.precision))

        ws = wb.create_sheet('概览')
        ws.append(['导出时间：'+datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
        header = ['学号','姓名','专业','学分绩','缺课门数','缺课表']
        ws.append(header)
        stu_grades = list(self.gradeLib.values())
        stu_grades.sort(key=lambda x:x.gpa,reverse=True)
        for stu_grade in stu_grades:
            stu_grade:StudentGrade
            ws.append(
                [str(stu_grade.student.number),stu_grade.student.name,stu_grade.student.major,
                 round(stu_grade.gpa,self.precision),stu_grade.absentCount]+stu_grade.absentNames
            )

        for major in CourseLib.ValidMajors:
            ws = wb.create_sheet(major.replace('*',''))
            # 学号，姓名，总分，课程
            courses = self.courseLib.majorCourseList(major,self.mode,self.start,self.end)
            header = ['学号','姓名','学分绩']+[course.name for course in courses]
            ws.append(header)
            second = ['','','']+[course.credits for course in courses]
            ws.append(second)
            major_list = []
            for stu_id,stu_grade in self.gradeLib.items():
                stu_grade:StudentGrade
                if stu_grade.student.major != major:
                    continue
                major_list.append(stu_grade)
            major_list.sort(key=lambda stu_grade:stu_grade.gpa,reverse=True)
            row = 2  # 当前行号
            start_col = 4
            for stu_grade in major_list:
                line = [str(stu_grade.student.number),stu_grade.student.name,
                        round(stu_grade.gpa,self.precision)]+\
                [stu_grade.getCourseGrade(course.id) for course in courses]
                ws.append(line)
                row += 1
                for c, course in enumerate(courses):
                    course:Course
                    grade = stu_grade.getCourseGradeObject(course.id)
                    txt = ""
                    if grade is None:
                        txt = "缺课，不计算本课程"
                    elif grade.note:
                        txt += f"Note1: {grade.note}\n"
                    elif grade.note2:
                        txt += f"Note2: {grade.note2}\n"
                    elif grade.flag:
                        txt += f"Flags: {grade.flag}\n"
                    if txt:
                        if grade is not None:
                            txt += f'修读学期：{grade.semester}'
                        ws.cell(row,c+start_col).comment = \
                            Comment(txt,'CeasGpaCalculator')


        wb.save(filename)
        wb.close()

    def saveIdOnlyExcel(self,filename='../data/output_id.xlsx'):
        """
        导出只包含学号的发布版文档。
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '概览'
        ws.append(['导出时间：' + datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
        header = ['学号','学分绩']
        ws.append(header)
        stu_grades = list(self.gradeLib.values())
        stu_grades.sort(key=lambda x: x.gpa, reverse=True)
        for stu_grade in stu_grades:
            stu_grade: StudentGrade
            ws.append(
                (str(stu_grade.student.number), round(stu_grade.gpa, self.precision))
            )

        for major in CourseLib.ValidMajors:
            ws = wb.create_sheet(major.replace('*',''))
            # 学号，姓名，总分，课程
            courses = self.courseLib.majorCourseList(major, self.mode, self.start, self.end)
            header = ('学号', '学分绩')
            ws.append(header)
            major_list = []  # 本专业学生表
            for stu_id, stu_grade in self.gradeLib.items():
                stu_grade: StudentGrade
                if stu_grade.student.major != major:
                    continue
                major_list.append(stu_grade)
            major_list.sort(key=lambda stu_grade: stu_grade.gpa, reverse=True)
            for stu_grade in major_list:
                line = (str(stu_grade.student.number), round(stu_grade.gpa, self.precision))
                ws.append(line)
        wb.save(filename)
        wb.close()